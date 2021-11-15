from datetime import datetime
from os.path import isfile
from openpyxl import load_workbook, Workbook


class ExperimentInfo:
    def __init__(self, name: str) -> None:
        self.name = name
        self.exp_count = 1
        self.try_count = 0
        self.record_time = 0
        self.last_reaction_time = 0


target_book: Workbook = None
if isfile('./result.xlsx'):
    target_book = load_workbook('./result.xlsx')
else:
    target_book = Workbook()


def record(info: ExperimentInfo):
    now = datetime.now()
    
    if info.name not in target_book.sheetnames:
        target_book.create_sheet(info.name, 0)
        sheet = target_book[info.name]
        sheet.append(['Name', 'Date Time', 'Experiment', 'Try', 'Time', 'Reaction Interval'])

    sheet = target_book[info.name]
    sheet.append([info.name, now.strftime('%Y-%m-%d %H:%M:%S'), info.exp_count, info.try_count, info.record_time, info.last_reaction_time / 1000000000])

    target_book.save('./result.xlsx')